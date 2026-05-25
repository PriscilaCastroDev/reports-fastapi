from datetime import datetime
from io import BytesIO

import openpyxl
import openpyxl.cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import settings

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _fmt(value: object) -> object:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def build_excel(rows: list[dict], columns: list[tuple[str, str]]) -> BytesIO:
    if not columns and rows:
        columns = [(k.replace("_", " ").title(), k) for k in rows[0].keys()]

    headers = [col[0] for col in columns]
    keys = [col[1] for col in columns]

    buffer = BytesIO()

    if len(rows) > settings.max_rows_sync:
        _build_streaming(buffer, rows, headers, keys)
    else:
        _build_normal(buffer, rows, headers, keys)

    buffer.seek(0)
    return buffer


def _build_normal(
    buffer: BytesIO,
    rows: list[dict],
    headers: list[str],
    keys: list[str],
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ws.append(headers)
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    col_widths = [len(h) for h in headers]

    for row in rows:
        values = [_fmt(row.get(k)) for k in keys]
        ws.append(values)
        for i, v in enumerate(values):
            col_widths[i] = max(col_widths[i], len(str(v)) if v is not None else 0)

    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 60)

    wb.save(buffer)


def _build_streaming(
    buffer: BytesIO,
    rows: list[dict],
    headers: list[str],
    keys: list[str],
) -> None:
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Reporte")

    header_row = []
    for h in headers:
        cell = openpyxl.cell.WriteOnlyCell(ws, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        header_row.append(cell)
    ws.append(header_row)

    for row in rows:
        ws.append([_fmt(row.get(k)) for k in keys])

    wb.save(buffer)
